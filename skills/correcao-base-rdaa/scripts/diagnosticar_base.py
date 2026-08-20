#!/usr/bin/env python3
"""
Diagnostico da base de contencioso (export CPJ-3C, planilha "Resolutivo" do RDAA).

Le a planilha exportada do CPJ-3C e gera um workbook de diagnostico com:
- Resumo Executivo (KPIs, alguns como formulas vivas)
- Plano de Acao consolidado e priorizado
- Abas de detalhe para cada categoria de achado
- Aba "Base (dados + auxiliares)" com os dados originais + colunas de apoio
  (formulas Excel) para o achado "recurso com origem arquivada"

Uso:
    python diagnosticar_base.py <entrada.xlsx> [saida.xlsx] [--sheet NOME_DA_ABA]

Se a saida nao for informada, usa "Diagnostico_e_Plano_de_Acao_<nome-base>.xlsx"
na mesma pasta da entrada.

IMPORTANTE: depois de gerar o arquivo, recalcule as formulas com o script
recalc.py da skill "xlsx" (LibreOffice) e confira que total_errors = 0 antes
de entregar o arquivo ao usuario. Este script sozinho NAO recalcula formulas.
"""
import sys
import re
import unicodedata
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Colunas esperadas no export CPJ-3C. Se o export mudar de layout, ajuste
# aqui -- o resto do script referencia colunas pelo NOME, nao pela posicao.
# ---------------------------------------------------------------------------
COL_CLIENTE = 'Cliente'
COL_NUMPROC = 'Número do processo'
COL_ACAO = 'Ação'
COL_AUTOR = 'Autor'
COL_REU = 'Réu'
COL_FASE = 'Fase Processual'
COL_RISCO = 'Risco'
COL_SITUACAO = 'Situação atual'
COL_RESUMO = 'Resumo/Assunto'
COL_ADVOGADO = 'Advogado Responsável'
COL_FICHA = 'Ficha'
COL_INCIDENTE = 'Arquivo Ficha Incidente'
COL_INSTANCIA = 'Instância'
COL_LOCALIZADOR = 'Localizador'  # este campo, apesar do nome, guarda o STATUS (ATIVO/ARQUIVADO/...)

REQUIRED_COLS = [COL_CLIENTE, COL_NUMPROC, COL_ACAO, COL_AUTOR, COL_REU, COL_FASE,
                 COL_RISCO, COL_SITUACAO, COL_RESUMO, COL_ADVOGADO, COL_FICHA,
                 COL_INCIDENTE, COL_INSTANCIA, COL_LOCALIZADOR]

# Padroes (normalizados, sem acento, minusculo) que classificam uma "Ação"
# como RECURSO no sentido juridico do termo (CPC art. 994 + praxe).
# Deliberadamente EXCLUIDOS (nao sao recurso, sao acao/incidente autonomo):
#   embargos a execucao, embargos de terceiro, impugnacao a credito habilitado,
#   conflito de competencia, incidente de desconsideracao de personalidade
#   juridica, acao rescisoria.
RECURSO_PATTERNS = [
    'agravo de instrumento', 'agravo interno', 'agravo regimental',
    'agravo em recurso especial', 'agravo recurso extraordinario',
    'agravo em recurso extraordinario', 'apelacao', 'recurso de apelacao',
    'embargos de declaracao', 'embargos declaracao', 'embargos infringentes',
    'recurso especial', 'recurso extraordinario', 'recurso inominado',
    'recurso ordinario', 'pedido de concessao de efeito suspensivo',
]

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
SUB_FONT = Font(name=FONT_NAME, bold=True, size=11, color='1F3864')
NORMAL_FONT = Font(name=FONT_NAME, size=10)
PRIORITY_FILL = PatternFill('solid', fgColor='FFC7CE')
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
INFO_FILL = PatternFill('solid', fgColor='DDEBF7')
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def norm_lower(s):
    if pd.isna(s):
        return ''
    s = str(s).lower().strip()
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()


def norm_nome(s):
    """Normaliza nomes de partes/clientes para comparacao (maiusculo, sem
    acento/pontuacao, sem sufixos societarios que atrapalham o match)."""
    if pd.isna(s):
        return ''
    s = str(s).upper().strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    s = re.sub(r'\b(S A|SA|LTDA|LTD|EIRELI|ME|EPP|JUNIOR|JR|FILHO|CIA|COMPANHIA)\b', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def is_recurso(acao):
    a = norm_lower(acao)
    return any(p in a for p in RECURSO_PATTERNS)


def is_arquivado_row(fase, localizador):
    f = str(fase).strip().upper() if pd.notna(fase) else ''
    l = str(localizador).strip().upper() if pd.notna(localizador) else ''
    return f.startswith('ARQUIVADO') or l.startswith('ARQUIVADO')


def qualidade_resumo(s):
    if pd.isna(s):
        return 'BRANCO'
    s = str(s).strip()
    if len(s) == 0:
        return 'BRANCO'
    if len(s) < 20:
        return 'MUITO CURTO / SEM CONTEUDO UTIL'
    if re.fullmatch(r'\[?recurso\]?\s*\S*', s, flags=re.IGNORECASE):
        return 'PLACEHOLDER (repete numero)'
    return 'OK'


def polo_cliente(cliente_n, autor_n, reu_n):
    if not cliente_n:
        return 'CLIENTE EM BRANCO'
    match_a = bool(autor_n) and (cliente_n in autor_n or autor_n in cliente_n)
    match_r = bool(reu_n) and (cliente_n in reu_n or reu_n in cliente_n)
    if match_a and match_r:
        return 'AMBIGUO (bate com autor e reu)'
    if match_a:
        return 'AUTOR'
    if match_r:
        return 'REU'
    return 'NAO IDENTIFICADO (nao bate com autor nem reu)'


# ---------------------------------------------------------------------------
# Analise
# ---------------------------------------------------------------------------

def carregar_e_analisar(caminho_entrada, sheet_name=None):
    xls = pd.ExcelFile(caminho_entrada)
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]
    df = pd.read_excel(caminho_entrada, sheet_name=sheet_name)

    faltando = [c for c in REQUIRED_COLS if c not in df.columns]
    if faltando:
        raise ValueError(
            f"Colunas esperadas nao encontradas no export: {faltando}. "
            f"Colunas disponiveis: {list(df.columns)}. "
            "Ajuste as constantes COL_* no topo do script se o layout do CPJ-3C mudou."
        )

    df['Sufixo'] = df[COL_INCIDENTE].astype(str).str.extract(r'\.(\d+)$')
    df['EhOrigem'] = df['Sufixo'] == '00'
    df['is_recurso'] = df[COL_ACAO].apply(is_recurso)
    df['EhArquivado'] = df.apply(lambda r: is_arquivado_row(r[COL_FASE], r[COL_LOCALIZADOR]), axis=1)
    df['Qualidade_Resumo'] = df[COL_RESUMO].apply(qualidade_resumo)

    df['Cliente_n'] = df[COL_CLIENTE].apply(norm_nome)
    df['Autor_n'] = df[COL_AUTOR].apply(norm_nome)
    df['Reu_n'] = df[COL_REU].apply(norm_nome)
    df['Polo_Cliente'] = df.apply(lambda r: polo_cliente(r['Cliente_n'], r['Autor_n'], r['Reu_n']), axis=1)

    # --- achado 1: recurso com origem sinalizada como arquivada (via Ficha .00) ---
    origem = df[df['EhOrigem']][[COL_FICHA, 'EhArquivado', COL_LOCALIZADOR, COL_FASE]].rename(
        columns={'EhArquivado': 'origem_arquivada', COL_LOCALIZADOR: 'origem_status', COL_FASE: 'origem_fase'})
    merged = df.merge(origem, on=COL_FICHA, how='left')
    prioritario = merged[(~merged['EhOrigem']) & (merged['origem_arquivada'] == True) & (merged['EhArquivado'] == False)].copy()

    # --- achado 1b: recurso "solto" -- sem origem rastreavel nem por numero nem por ficha ---
    num_groups = df.groupby(COL_NUMPROC)
    ficha_groups = df.groupby(COL_FICHA)

    def tem_origem_visivel(row, key_col, groups):
        key = row[key_col]
        if pd.isna(key):
            return False
        grupo = groups.get_group(key)
        outros = grupo[grupo.index != row.name]
        if len(outros) == 0:
            return False
        tem_acao_origem = (~outros['is_recurso']).any()
        tem_instancia1 = (outros[COL_INSTANCIA].astype(str).str.strip() == '1').any()
        return bool(tem_acao_origem or tem_instancia1)

    df['origem_por_numero'] = df.apply(lambda r: tem_origem_visivel(r, COL_NUMPROC, num_groups), axis=1)
    df['origem_por_ficha'] = df.apply(lambda r: tem_origem_visivel(r, COL_FICHA, ficha_groups), axis=1)
    df['tem_origem_rastreavel'] = df['origem_por_numero'] | df['origem_por_ficha']
    df['RECURSO_SOLTO'] = df['is_recurso'] & (~df['tem_origem_rastreavel'])
    solto_out = df[df['RECURSO_SOLTO']][[COL_FICHA, COL_NUMPROC, COL_ACAO, COL_INSTANCIA, COL_LOCALIZADOR,
                                          COL_FASE, COL_CLIENTE, COL_ADVOGADO]].sort_values(COL_FICHA)

    # --- achado 2: inconsistencia Fase Processual x Localizador (status) ---
    fase_arq = df[COL_FASE].astype(str).str.strip().str.upper().str.startswith('ARQUIVADO')
    loc_arq = df[COL_LOCALIZADOR].astype(str).str.strip().str.upper().eq('ARQUIVADO')
    inconsist = df[fase_arq & ~loc_arq].copy()
    inconsist_out = inconsist[[COL_FICHA, COL_INCIDENTE, COL_CLIENTE, COL_NUMPROC, COL_ACAO,
                                COL_LOCALIZADOR, COL_FASE, 'EhOrigem', COL_ADVOGADO]].rename(
        columns={COL_LOCALIZADOR: 'Status atual (campo)', COL_FASE: 'Fase Processual (campo)',
                 'EhOrigem': 'É registro de origem?'})
    inconsist_out['Recomendação'] = 'Confirmar status real do processo e uniformizar os campos "Status" e "Fase Processual" (hoje divergentes).'

    # --- achado 3: registros de teste ---
    testes = df[df[COL_NUMPROC].astype(str).str.strip().str.lower().isin(['teste', 'testes'])]
    testes_out = testes[[COL_FICHA, COL_INCIDENTE, COL_NUMPROC, COL_ACAO, COL_CLIENTE, COL_ADVOGADO]].copy()
    testes_out['Recomendação'] = 'Excluir registro de teste da base — não é processo real.'

    # --- achado 4: acao/cliente/polo a revisar ---
    prob_ac = df[(df[COL_ACAO].isna()) | (df['Polo_Cliente'].isin(
        ['NAO IDENTIFICADO (nao bate com autor nem reu)', 'AMBIGUO (bate com autor e reu)', 'CLIENTE EM BRANCO']))]
    prob_ac_out = prob_ac[[COL_FICHA, COL_NUMPROC, COL_ACAO, COL_CLIENTE, COL_AUTOR, COL_REU, 'Polo_Cliente', COL_ADVOGADO]].copy()
    prob_ac_out['Observação'] = prob_ac_out['Polo_Cliente'].map({
        'CLIENTE EM BRANCO': 'Campo Cliente vazio — preencher.',
        'NAO IDENTIFICADO (nao bate com autor nem reu)': 'Cliente não corresponde a Autor nem a Réu — verificar se é terceiro interessado/endossatário ou erro de cadastro.',
        'AMBIGUO (bate com autor e reu)': 'Nome do cliente coincide com Autor e Réu — checar homônimo ou erro de cadastro.',
    }).fillna('Ação (tipo de peça) não preenchida.')
    prob_ac_out = prob_ac_out.sort_values(COL_FICHA)

    # --- achado 5: qualidade do resumo ---
    prob_resumo = df[df['Qualidade_Resumo'] != 'OK'][[COL_FICHA, COL_NUMPROC, COL_ACAO, COL_RESUMO,
                                                        'Qualidade_Resumo', COL_ADVOGADO]].sort_values(COL_FICHA)

    # --- achado 6: campos criticos em branco, por advogado ---
    def blank_pivot(col):
        b = df[df[col].isna()]
        return b[COL_ADVOGADO].fillna('(sem advogado)').value_counts()

    pivot_df = pd.DataFrame({
        'Sem Risco classificado': blank_pivot(COL_RISCO),
        'Sem Situação atual': blank_pivot(COL_SITUACAO),
        'Sem Fase Processual': blank_pivot(COL_FASE),
    }).fillna(0).astype(int)
    pivot_df.index.name = COL_ADVOGADO
    pivot_df = pivot_df.reset_index()
    if len(pivot_df):
        pivot_df.loc['TOTAL'] = ['TOTAL'] + pivot_df.iloc[:, 1:].sum().tolist()

    return {
        'df': df, 'n': len(df), 'sheet_name': sheet_name,
        'prioritario': prior_cols(prioritario, COL_FICHA, COL_INCIDENTE, COL_CLIENTE, COL_NUMPROC, COL_ACAO,
                                   COL_LOCALIZADOR, COL_FASE, COL_ADVOGADO),
        'solto': solto_out, 'inconsist': inconsist_out, 'testes': testes_out,
        'prob_ac': prob_ac_out, 'prob_resumo': prob_resumo, 'pivot': pivot_df,
    }


def prior_cols(prioritario, ficha, incidente, cliente, numproc, acao, localizador, fase, advogado):
    cols = [ficha, incidente, cliente, 'origem_status', numproc, acao, localizador, fase, 'origem_status', 'origem_fase', advogado]
    out = prioritario[[ficha, incidente, cliente, numproc, acao, localizador, fase, 'origem_status', 'origem_fase', advogado]].rename(
        columns={localizador: 'Status do Recurso', fase: 'Fase do Recurso',
                 'origem_status': 'Status da Origem', 'origem_fase': 'Fase da Origem'})
    out.insert(0, 'Prioridade', 'ALTA')
    out['Recomendação'] = ('Confirmar no tribunal se o recurso já foi julgado (origem consta arquivada). '
                            'Se julgado, atualizar status/baixa e encerrar provisão. Se não, corrigir o campo '
                            '"Fase Processual" da origem, que está incorreto.')
    return out


# ---------------------------------------------------------------------------
# Escrita do workbook
# ---------------------------------------------------------------------------

def write_df(ws, df_, start_row=1, table_name=None):
    for j, c_ in enumerate(df_.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(c_))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(df_.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = ''
            c = ws.cell(row=i, column=j, value=val)
            c.font = NORMAL_FONT
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
    end_row = start_row + len(df_)
    for j, colname in enumerate(df_.columns, start=1):
        maxlen = max([len(str(colname))] + [len(str(v)) for v in df_[colname].astype(str)]) if len(df_) else len(str(colname))
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, maxlen + 2), 45)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    ws.sheet_view.showGridLines = False
    if table_name and len(df_) > 0:
        ref = f"A{start_row}:{get_column_letter(len(df_.columns))}{end_row}"
        tbl = Table(displayName=table_name[:30], ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    return end_row


def montar_workbook(resultado, caminho_saida):
    df = resultado['df']
    n = resultado['n']
    lastrow = n + 1

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- aba raw + auxiliares (com formulas para o achado 1) ----------------
    ws_raw = wb.create_sheet('Base (dados + auxiliares)')
    df_export_cols = [c for c in df.columns if c not in (
        'Sufixo', 'EhOrigem', 'is_recurso', 'EhArquivado', 'Qualidade_Resumo', 'Cliente_n', 'Autor_n', 'Reu_n',
        'Polo_Cliente', 'origem_por_numero', 'origem_por_ficha', 'tem_origem_rastreavel', 'RECURSO_SOLTO')]
    df_raw = df[df_export_cols]
    write_df(ws_raw, df_raw, start_row=1)
    ncols = len(df_export_cols)
    first_aux_col = ncols + 1

    aux_headers = ['FaseIndicaArquivado', 'LocalizadorArquivado', 'EhArquivado(qualquer sinal)',
                   'Sufixo(Ficha Incidente)', 'EhOrigem(.00)', 'ChaveOrigemFicha', 'ChavePropriaFicha',
                   'OrigemDaFichaEstaArquivada', 'RECURSO C/ ORIGEM ARQUIVADA (revisar)',
                   'Fase diz Arquivado mas Status nao (inconsistencia)']
    for j, h in enumerate(aux_headers, start=first_aux_col):
        c = ws_raw.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = PatternFill('solid', fgColor='548235')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

    def col_letter(name):
        return get_column_letter(df_export_cols.index(name) + 1)

    C_FASE, C_LOC, C_FICHA_, C_INC = (col_letter(COL_FASE), col_letter(COL_LOCALIZADOR),
                                       col_letter(COL_FICHA), col_letter(COL_INCIDENTE))
    L = [get_column_letter(first_aux_col + k) for k in range(10)]
    FASE_OK, LOC_OK, EH_ARQ, SUFIXO, EH_ORIGEM, CHV_ORIG, CHV_PROP, ORIG_ARQ, PRIORIDADE, INCONSIST = L

    for r in range(2, lastrow + 1):
        ws_raw[f'{FASE_OK}{r}'] = f'=LEFT(UPPER(TRIM({C_FASE}{r})),9)="ARQUIVADO"'
        ws_raw[f'{LOC_OK}{r}'] = f'=UPPER(TRIM({C_LOC}{r}))="ARQUIVADO"'
        ws_raw[f'{EH_ARQ}{r}'] = f'=OR({FASE_OK}{r},{LOC_OK}{r})'
        ws_raw[f'{SUFIXO}{r}'] = f'=IFERROR(RIGHT({C_INC}{r},LEN({C_INC}{r})-FIND(".",{C_INC}{r})),"")'
        ws_raw[f'{EH_ORIGEM}{r}'] = f'={SUFIXO}{r}="00"'
        ws_raw[f'{CHV_ORIG}{r}'] = f'={C_FICHA_}{r}&"|00"'
        ws_raw[f'{CHV_PROP}{r}'] = f'={C_FICHA_}{r}&"|"&{SUFIXO}{r}'
        ws_raw[f'{ORIG_ARQ}{r}'] = (f'=IFERROR(INDEX(${EH_ARQ}$2:${EH_ARQ}${lastrow},'
                                     f'MATCH({CHV_ORIG}{r},${CHV_PROP}$2:${CHV_PROP}${lastrow},0)),FALSE)')
        ws_raw[f'{PRIORIDADE}{r}'] = f'=AND(NOT({EH_ORIGEM}{r}),{ORIG_ARQ}{r},NOT({EH_ARQ}{r}))'
        ws_raw[f'{INCONSIST}{r}'] = f'=AND({FASE_OK}{r},NOT({LOC_OK}{r}))'
        for cc in L:
            ws_raw[f'{cc}{r}'].font = NORMAL_FONT
            ws_raw[f'{cc}{r}'].border = BORDER

    for j in range(1, ncols + 1):
        ws_raw.column_dimensions[get_column_letter(j)].width = 14
    for cc in L:
        ws_raw.column_dimensions[cc].width = 16
    ws_raw.auto_filter.ref = f"A1:{L[-1]}{lastrow}"
    ws_raw.sheet_view.showGridLines = False

    # ---------------- abas de detalhe ----------------
    def add_sheet(name, df_, priority_fill_all=False):
        ws = wb.create_sheet(name)
        end_row = write_df(ws, df_, 1, table_name=name.replace(' ', '_')[:28])
        if priority_fill_all:
            for r in range(2, end_row + 1):
                for c in range(1, df_.shape[1] + 1):
                    ws.cell(row=r, column=c).fill = PRIORITY_FILL
        return ws

    add_sheet('1. Recursos c Origem Arquivada', resultado['prioritario'], priority_fill_all=True)
    add_sheet('1b. Recursos Soltos', resultado['solto'], priority_fill_all=True)
    add_sheet('2. Inconsistencia Status-Fase', resultado['inconsist'])
    add_sheet('3. Registros de Teste', resultado['testes'])
    add_sheet('5. Campos Criticos em Branco', resultado['pivot'])
    add_sheet('6. Acao-Cliente-Polo', resultado['prob_ac'])
    add_sheet('7. Qualidade do Resumo', resultado['prob_resumo'])

    # ---------------- Plano de Acao ----------------
    rows = []
    for _, r in resultado['prioritario'].iterrows():
        rows.append(['1 - Alta', 'Recurso c/ origem arquivada', r[COL_FICHA], r[COL_NUMPROC], r[COL_ACAO],
                     r[COL_CLIENTE], r[COL_ADVOGADO], r['Recomendação'], '5 dias úteis'])
    for _, r in resultado['solto'].iterrows():
        rows.append(['1 - Alta', 'Recurso solto (sem origem rastreável)', r[COL_FICHA], r[COL_NUMPROC], r[COL_ACAO],
                     r[COL_CLIENTE], r[COL_ADVOGADO],
                     'Localizar/confirmar o processo de origem no tribunal (provável arquivamento/baixa). '
                     'Vincular à ficha de origem ou registrar a informação de origem na ficha do recurso.',
                     '10 dias úteis'])
    for _, r in resultado['inconsist'].iterrows():
        rows.append(['2 - Média', 'Status x Fase Processual divergentes', r[COL_FICHA], r[COL_NUMPROC], r[COL_ACAO],
                     r[COL_CLIENTE], r[COL_ADVOGADO], r['Recomendação'], '15 dias úteis'])
    for _, r in resultado['prob_ac'].iterrows():
        rows.append(['2 - Média', 'Ação/Cliente/Polo a revisar', r[COL_FICHA], r[COL_NUMPROC], r[COL_ACAO],
                     r[COL_CLIENTE], r[COL_ADVOGADO], r['Observação'], '15 dias úteis'])
    for _, r in resultado['testes'].iterrows():
        rows.append(['3 - Baixa', 'Registro de teste', r[COL_FICHA], r[COL_NUMPROC], r[COL_ACAO],
                     r[COL_CLIENTE], r[COL_ADVOGADO], r['Recomendação'], '2 dias úteis'])

    plano_df = pd.DataFrame(rows, columns=['Prioridade', 'Categoria', 'Ficha', 'Nº Processo', 'Ação', 'Cliente',
                                            'Advogado Responsável', 'Ação recomendada', 'Prazo sugerido'])
    plano_df = plano_df.sort_values('Prioridade')

    ws_p = wb.create_sheet('Plano de Ação', 1)
    for j, c_ in enumerate(plano_df.columns, start=1):
        c = ws_p.cell(row=1, column=j, value=c_)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    fill_map = {'1 - Alta': PRIORITY_FILL, '2 - Média': WARN_FILL, '3 - Baixa': INFO_FILL}
    for i, row in enumerate(plano_df.itertuples(index=False), start=2):
        fill = fill_map.get(row[0])
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = ''
            c = ws_p.cell(row=i, column=j, value=val)
            c.font = NORMAL_FONT
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
            if fill:
                c.fill = fill
    end_row = len(plano_df) + 1
    widths = [10, 26, 8, 22, 22, 28, 24, 55, 14]
    for j, w in enumerate(widths, start=1):
        ws_p.column_dimensions[get_column_letter(j)].width = w
    ws_p.freeze_panes = 'A2'
    ws_p.sheet_view.showGridLines = False
    if len(plano_df) > 0:
        tbl = Table(displayName="PlanoDeAcao", ref=f"A1:{get_column_letter(len(plano_df.columns))}{end_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws_p.add_table(tbl)

    # ---------------- Resumo Executivo ----------------
    ws_r = wb.create_sheet('Resumo Executivo', 0)
    ws_r.sheet_view.showGridLines = False
    ws_r['B2'] = 'Diagnóstico da Base (CPJ-3C) — RDAA'
    ws_r['B2'].font = Font(name=FONT_NAME, bold=True, size=16, color='1F3864')
    ws_r['B3'] = f'Base analisada: {Path(caminho_saida).stem} | Aba de origem: {resultado["sheet_name"]}'
    ws_r['B3'].font = Font(name=FONT_NAME, italic=True, size=10, color='595959')

    def kpi(row, label, formula, fill=None):
        ws_r.cell(row=row, column=2, value=label).font = Font(name=FONT_NAME, size=10)
        c = ws_r.cell(row=row, column=5, value=formula)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color='1F3864')
        c.alignment = Alignment(horizontal='right')
        if fill:
            c.fill = fill

    R = "'Base (dados + auxiliares)'"
    Z = col_letter(COL_LOCALIZADOR)
    W = col_letter(COL_FICHA)
    O_ = col_letter(COL_RISCO)
    Q_ = col_letter(COL_SITUACAO)
    H_ = col_letter(COL_FASE)
    B_ = col_letter(COL_CLIENTE)
    D_ = col_letter(COL_ACAO)

    row = 5
    ws_r.cell(row=row, column=2, value='INDICADORES GERAIS').font = SUB_FONT; row += 1
    kpi(row, 'Total de registros na base', f'=COUNTA({R}!${{}}$2:${{}}${{}})'.format(get_column_letter(1), get_column_letter(1), lastrow)); row += 1
    kpi(row, 'Total de fichas (casos/incidentes) cadastradas', f'=SUMPRODUCT(1/COUNTIF({R}!${W}$2:${W}${lastrow},{R}!${W}$2:${W}${lastrow}))'); row += 1
    kpi(row, 'Registros classificados juridicamente como RECURSO', int(df['is_recurso'].sum())); row += 1
    row += 1

    ws_r.cell(row=row, column=2, value='PRIORIDADE 1 — RECURSOS COM ORIGEM ARQUIVADA (sinalizada)').font = SUB_FONT; row += 1
    kpi(row, 'Recursos ativos cuja ficha de origem já consta "Arquivado"', f'=COUNTIF({R}!${PRIORIDADE}$2:${PRIORIDADE}${lastrow},TRUE)', PRIORITY_FILL); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "1. Recursos c Origem Arquivada"').font = Font(name=FONT_NAME, italic=True, size=9, color='C00000'); row += 2

    ws_r.cell(row=row, column=2, value='PRIORIDADE 1 — RECURSOS SOLTOS (sem processo de origem rastreável na base)').font = SUB_FONT; row += 1
    kpi(row, 'Recursos sem nenhum registro de origem (mesmo nº processo/ficha, ação de origem ou instância=1)', int(df['RECURSO_SOLTO'].sum()), PRIORITY_FILL); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "1b. Recursos Soltos" — provável processo de origem arquivado/baixado e não rastreado').font = Font(name=FONT_NAME, italic=True, size=9, color='C00000'); row += 2

    ws_r.cell(row=row, column=2, value='PRIORIDADE 2 — INCONSISTÊNCIAS DE CADASTRO').font = SUB_FONT; row += 1
    kpi(row, 'Fase Processual = "Arquivado" mas Status (Localizador) diferente', f'=COUNTIF({R}!${INCONSIST}$2:${INCONSIST}${lastrow},TRUE)', WARN_FILL); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "2. Inconsistencia Status x Fase"').font = Font(name=FONT_NAME, italic=True, size=9, color='9C6500'); row += 2

    ws_r.cell(row=row, column=2, value='PRIORIDADE 2 — AÇÃO / CLIENTE / POLO').font = SUB_FONT; row += 1
    kpi(row, 'Registros sem Cliente preenchido', f'=COUNTBLANK({R}!${B_}$2:${B_}${lastrow})'); row += 1
    kpi(row, 'Registros sem Ação (tipo de peça) preenchida', f'=COUNTBLANK({R}!${D_}$2:${D_}${lastrow})'); row += 1
    kpi(row, 'Cliente não corresponde a Autor nem a Réu (polo não identificado/ambíguo)', len(resultado['prob_ac']), WARN_FILL); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "6. Acao-Cliente-Polo"').font = Font(name=FONT_NAME, italic=True, size=9, color='9C6500'); row += 2

    ws_r.cell(row=row, column=2, value='PRIORIDADE 2 — QUALIDADE DO RESUMO/ASSUNTO').font = SUB_FONT; row += 1
    kpi(row, 'Resumo/Assunto em branco ou sem conteúdo útil (curto/placeholder)', len(resultado['prob_resumo']), WARN_FILL); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "7. Qualidade do Resumo"').font = Font(name=FONT_NAME, italic=True, size=9, color='9C6500'); row += 2

    ws_r.cell(row=row, column=2, value='PRIORIDADE 3 — CAMPOS CRÍTICOS EM BRANCO (mutirão)').font = SUB_FONT; row += 1
    kpi(row, 'Registros sem classificação de Risco', f'=COUNTBLANK({R}!${O_}$2:${O_}${lastrow})'); row += 1
    kpi(row, 'Registros sem "Situação atual" preenchida', f'=COUNTBLANK({R}!${Q_}$2:${Q_}${lastrow})'); row += 1
    kpi(row, 'Registros sem "Fase Processual" preenchida', f'=COUNTBLANK({R}!${H_}$2:${H_}${lastrow})'); row += 1
    ws_r.cell(row=row, column=2, value='→ Ver aba "5. Campos Criticos em Branco"').font = Font(name=FONT_NAME, italic=True, size=9, color='595959'); row += 2

    ws_r.cell(row=row, column=2, value='OUTROS ACHADOS').font = SUB_FONT; row += 1
    kpi(row, 'Registros de teste a remover da base', len(resultado['testes'])); row += 1
    ws_r.cell(row=row, column=2, value='Nota: números de processo repetidos em fichas diferentes NÃO são duplicidade —').font = Font(name=FONT_NAME, italic=True, size=9, color='595959'); row += 1
    ws_r.cell(row=row, column=2, value='é a prática normal do escritório de abrir nova ficha a cada recurso interposto.').font = Font(name=FONT_NAME, italic=True, size=9, color='595959'); row += 2

    ws_r.cell(row=row, column=2, value='Como usar este arquivo').font = SUB_FONT; row += 1
    notas = [
        '1. A aba "Base (dados + auxiliares)" traz os dados originais + colunas de apoio com fórmulas vivas para o',
        '   achado "origem arquivada" (recalculam se você colar uma nova exportação do CPJ-3C, mantendo o cabeçalho).',
        '2. As classificações mais elaboradas (recurso solto, polo do cliente, qualidade do resumo) exigem rodar',
        '   novamente este script sobre uma nova exportação — não recalculam sozinhas dentro do Excel.',
        '3. A aba "Plano de Ação" consolida os achados priorizados, com responsável e prazo sugeridos.',
        '4. Depois de gerar este arquivo, rode scripts/recalc.py (skill xlsx) para recalcular fórmulas e confira',
        '   que total_errors = 0 antes de entregar.',
    ]
    for nline in notas:
        ws_r.cell(row=row, column=2, value=nline).font = Font(name=FONT_NAME, size=9, italic=True, color='595959'); row += 1

    ws_r.column_dimensions['A'].width = 2
    ws_r.column_dimensions['B'].width = 78
    ws_r.column_dimensions['E'].width = 14

    order = ['Resumo Executivo', 'Plano de Ação', '1. Recursos c Origem Arquivada', '1b. Recursos Soltos',
             '2. Inconsistencia Status-Fase', '6. Acao-Cliente-Polo', '7. Qualidade do Resumo',
             '3. Registros de Teste', '5. Campos Criticos em Branco', 'Base (dados + auxiliares)']
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]

    wb.save(caminho_saida)
    return caminho_saida


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('entrada', help='Caminho do xlsx exportado do CPJ-3C')
    ap.add_argument('saida', nargs='?', default=None, help='Caminho do xlsx de diagnostico a gerar')
    ap.add_argument('--sheet', default=None, help='Nome da aba a ler (default: primeira aba)')
    args = ap.parse_args()

    entrada = Path(args.entrada)
    saida = Path(args.saida) if args.saida else entrada.parent / f"Diagnostico_e_Plano_de_Acao_{entrada.stem}.xlsx"

    resultado = carregar_e_analisar(entrada, sheet_name=args.sheet)
    montar_workbook(resultado, saida)

    print(f"OK: {resultado['n']} registros analisados.")
    print(f"Recursos: {int(resultado['df']['is_recurso'].sum())}")
    print(f"Recursos com origem arquivada (sinalizada): {len(resultado['prioritario'])}")
    print(f"Recursos soltos (sem origem rastreavel): {len(resultado['solto'])}")
    print(f"Inconsistencias status x fase: {len(resultado['inconsist'])}")
    print(f"Registros de teste: {len(resultado['testes'])}")
    print(f"Acao/Cliente/Polo a revisar: {len(resultado['prob_ac'])}")
    print(f"Resumo/Assunto de baixa qualidade: {len(resultado['prob_resumo'])}")
    print(f"Arquivo gerado: {saida}")
    print("PROXIMO PASSO OBRIGATORIO: rode o recalc.py da skill xlsx neste arquivo e confirme total_errors=0.")


if __name__ == '__main__':
    main()
