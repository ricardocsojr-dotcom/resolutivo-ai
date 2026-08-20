#!/usr/bin/env python3
"""
Monta a planilha final de analise de provisao/risco (com double-check) a
partir de um CSV com uma linha por processo.

O CSV e' o resultado do raciocinio juridico feito pelo Claude ao aplicar a
metodologia em references/metodologia-provisao.md (arvore b.1-b.7 + criterio
de auditoria CPC 25) processo a processo. Este script NAO classifica nada
sozinho -- ele so formata e consolida o que ja foi decidido, calcula os
totais por classificacao via formulas do Excel, e destaca divergencias.

Colunas esperadas no CSV de entrada (use exatamente estes nomes de cabecalho):

  Numero do processo, Ficha, Cliente, Partes, Natureza da demanda,
  Objeto resumido, Fase processual, Ultimo andamento relevante,
  Pedidos com impacto economico, Valor da causa, Valor economico atualizado,
  Depositos garantias ou bloqueios, Proximos prazos ou eventos,
  Data-base da analise, Responsavel pela avaliacao,
  Classificacao sugerida (usuario), Valor sugerido (usuario),
  Justificativa (usuario),
  Classificacao RDAA (double-check), Fundamento (double-check),
  Valor de contingencia (double-check), Valor provisionavel (double-check),
  Providencia recomendada, Divergencia, Limitacao ou observacao

Linhas com informacao insuficiente para estimar valor devem trazer a
limitacao explicita na coluna "Limitacao ou observacao" -- NUNCA um valor
arbitrario nas colunas de valor.

Uso:
    python montar_planilha_provisao.py <entrada.csv> <saida.xlsx>
"""
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = 'Arial'
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
SUB_FONT = Font(name=FONT_NAME, bold=True, size=11, color='1F3864')
NORMAL_FONT = Font(name=FONT_NAME, size=10)
DIVERGENCIA_FILL = PatternFill('solid', fgColor='FFC7CE')
RISCO_FILL = {
    'PROVAVEL': PatternFill('solid', fgColor='F8CBAD'),
    'POSSIVEL': PatternFill('solid', fgColor='FFE699'),
    'REMOTO': PatternFill('solid', fgColor='C6E0B4'),
}
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_CLASS_RDAA = 'Classificacao RDAA (double-check)'
COL_DIVERGENCIA = 'Divergencia'
COL_VALOR_PROVISIONAVEL = 'Valor provisionavel (double-check)'
COL_VALOR_CONTINGENCIA = 'Valor de contingencia (double-check)'

VALUE_COLS = ['Valor da causa', 'Valor economico atualizado', 'Valor sugerido (usuario)',
              COL_VALOR_CONTINGENCIA, COL_VALOR_PROVISIONAVEL]


def norm_risco(s):
    if pd.isna(s):
        return ''
    s = str(s).strip().upper()
    for a, b in (('Í', 'I'), ('Á', 'A'), ('Â', 'A'), ('É', 'E')):
        s = s.replace(a, b)
    return s


def to_number_if_plain(val):
    if pd.isna(val):
        return val
    s = str(val).strip()
    if s == '':
        return val
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return val


def write_df(ws, df_, start_row=1, table_name=None):
    for j, c_ in enumerate(df_.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(c_))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(df_.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = ''
            c = ws.cell(row=i, column=j, value=val)
            c.font = NORMAL_FONT
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
    end_row = start_row + len(df_)
    for j, colname in enumerate(df_.columns, start=1):
        maxlen = max([len(str(colname))] + [len(str(v)) for v in df_[colname].astype(str)]) if len(df_) else len(str(colname))
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, maxlen + 2), 45)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    ws.sheet_view.showGridLines = False
    if table_name and len(df_) > 0:
        ref = "A{}:{}{}".format(start_row, get_column_letter(len(df_.columns)), end_row)
        tbl = Table(displayName=table_name[:30], ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    return end_row


def montar(caminho_csv, caminho_saida):
    df = pd.read_csv(caminho_csv, dtype=str)
    for vc in VALUE_COLS:
        if vc in df.columns:
            df[vc] = df[vc].apply(to_number_if_plain)
    if COL_CLASS_RDAA not in df.columns:
        raise ValueError(
            "Coluna obrigatoria '{}' nao encontrada no CSV. Colunas presentes: {}".format(
                COL_CLASS_RDAA, list(df.columns))
        )

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Analise de Provisao', 0)
    end_row = write_df(ws, df, 1, table_name='AnaliseProvisao')

    class_col_idx = df.columns.get_loc(COL_CLASS_RDAA) + 1
    diverg_col_idx = df.columns.get_loc(COL_DIVERGENCIA) + 1 if COL_DIVERGENCIA in df.columns else None

    for r in range(2, end_row + 1):
        classe = norm_risco(ws.cell(row=r, column=class_col_idx).value)
        fill = RISCO_FILL.get(classe)
        if fill:
            ws.cell(row=r, column=class_col_idx).fill = fill
        if diverg_col_idx:
            dv = str(ws.cell(row=r, column=diverg_col_idx).value or '').strip().upper()
            if dv.startswith('SIM'):
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r, column=c).fill = DIVERGENCIA_FILL

    ws_r = wb.create_sheet('Resumo', 0)
    ws_r.sheet_view.showGridLines = False
    ws_r['B2'] = 'Analise de Provisao - Double-Check (RDAA)'
    ws_r['B2'].font = Font(name=FONT_NAME, bold=True, size=16, color='1F3864')
    ws_r['B3'] = 'Metodologia: arvore institucional (b.1-b.7) + criterio de auditoria CPC 25/NBC TG 25'
    ws_r['B3'].font = Font(name=FONT_NAME, italic=True, size=10, color='595959')

    R = "'Analise de Provisao'"
    CL = get_column_letter(class_col_idx)
    lastrow = end_row

    def kpi(row, label, formula, fill=None):
        ws_r.cell(row=row, column=2, value=label).font = Font(name=FONT_NAME, size=10)
        c = ws_r.cell(row=row, column=5, value=formula)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color='1F3864')
        c.alignment = Alignment(horizontal='right')
        if fill:
            c.fill = fill

    row = 5
    ws_r.cell(row=row, column=2, value='CONTAGEM POR CLASSIFICACAO (double-check RDAA)').font = SUB_FONT
    row += 1
    kpi(row, 'Provavel', '=COUNTIF({}!${}$2:${}${},"Prov*")'.format(R, CL, CL, lastrow), RISCO_FILL['PROVAVEL'])
    row += 1
    kpi(row, 'Possivel', '=COUNTIF({}!${}$2:${}${},"Poss*")'.format(R, CL, CL, lastrow), RISCO_FILL['POSSIVEL'])
    row += 1
    kpi(row, 'Remoto', '=COUNTIF({}!${}$2:${}${},"Remot*")'.format(R, CL, CL, lastrow), RISCO_FILL['REMOTO'])
    row += 2

    if COL_VALOR_PROVISIONAVEL in df.columns:
        VP = get_column_letter(df.columns.get_loc(COL_VALOR_PROVISIONAVEL) + 1)
        ws_r.cell(row=row, column=2, value='VALOR PROVISIONAVEL (double-check, soma dos numericos preenchidos)').font = SUB_FONT
        row += 1
        kpi(row, 'Total', '=SUM({}!${}$2:${}${})'.format(R, VP, VP, lastrow))
        row += 2

    if diverg_col_idx:
        DV = get_column_letter(diverg_col_idx)
        ws_r.cell(row=row, column=2, value='DIVERGENCIAS ENTRE SUGESTAO DO USUARIO E DOUBLE-CHECK').font = SUB_FONT
        row += 1
        kpi(row, 'Processos com divergencia (revisar prioritariamente)',
            '=COUNTIF({}!${}$2:${}${},"Sim*")'.format(R, DV, DV, lastrow), DIVERGENCIA_FILL)
        row += 2

    if 'Limitacao ou observacao' in df.columns:
        LI = get_column_letter(df.columns.get_loc('Limitacao ou observacao') + 1)
        ws_r.cell(row=row, column=2, value='LIMITACOES DECLARADAS (sem estimativa segura de valor)').font = SUB_FONT
        row += 1
        kpi(row, 'Processos com limitacao registrada', '=COUNTIF({}!${}$2:${}${},"?*")'.format(R, LI, LI, lastrow))
        row += 2

    ws_r.cell(row=row, column=2, value='Como usar').font = SUB_FONT
    row += 1
    notas = [
        '1. A aba "Analise de Provisao" tem uma linha por processo com os campos minimos + a comparacao',
        '   entre a classificacao sugerida pelo usuario e o double-check aplicado conforme a metodologia RDAA.',
        '2. Linhas destacadas em vermelho-claro tem divergencia entre a sugestao do usuario e o double-check --',
        '   comece a revisao por elas.',
        '3. A cor de cada celula de classificacao (Provavel/Possivel/Remoto) segue o double-check, nao a sugestao',
        '   original do usuario.',
        '4. Nunca ha valor arbitrario nas colunas financeiras: quando a estimativa nao e segura, a coluna',
        '   "Limitacao ou observacao" explica o motivo e o que falta para mensurar.',
        '5. A classificacao contabil final (reconhecer provisao, divulgar em nota, ou apenas monitorar) e do',
        '   cliente/contabilidade conforme CPC 25/NBC TG 25 -- este relatorio entrega o insumo juridico, nao a decisao contabil.',
    ]
    for nline in notas:
        ws_r.cell(row=row, column=2, value=nline).font = Font(name=FONT_NAME, size=9, italic=True, color='595959')
        row += 1

    ws_r.column_dimensions['A'].width = 2
    ws_r.column_dimensions['B'].width = 90
    ws_r.column_dimensions['E'].width = 14

    wb.save(caminho_saida)
    return caminho_saida


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('entrada', help='CSV com uma linha por processo (ver cabecalho esperado no topo deste arquivo)')
    ap.add_argument('saida', help='Caminho do xlsx final a gerar')
    args = ap.parse_args()
    caminho = montar(args.entrada, args.saida)
    print("OK: planilha gerada em {}".format(caminho))
    print("PROXIMO PASSO OBRIGATORIO: rode o recalc.py da skill xlsx neste arquivo e confirme total_errors=0.")


if __name__ == '__main__':
    main()
