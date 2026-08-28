#!/usr/bin/env python3
"""Self-check do renderizador simples. Rodar: python test_renderizar_memoria_simples.py"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from renderizar_memoria_simples import RenderError, render

TEMPLATE = Path(__file__).resolve().parents[1] / "references" / "template-calculo-simples-rdaa.xlsx"


def run() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        output = Path(tmp) / "saida.xlsx"
        payload = {
            "parcelas": [
                {"tipo": "Principal", "mes_ref": "01/2024", "vencimento": "10/01/2024", "valor": "100.00", "subtotal": "110.00"},
                {"tipo": "Honorários", "mes_ref": "", "vencimento": "", "valor": "10.00", "subtotal": "10.00"},
            ],
            "notas": ["Racional do teste", "", "Segunda linha"],
        }
        total = render(payload, template=TEMPLATE, output=output)
        assert total == 2, total

        wb = load_workbook(output, data_only=False)
        ws = wb["Cálculo"]
        assert ws.cell(2, 1).value == "Principal", ws.cell(2, 1).value
        assert ws.cell(3, 1).value == "Honorários", ws.cell(3, 1).value
        assert ws.cell(2, 4).value == 100.00, ws.cell(2, 4).value
        assert ws.cell(501, 2).value == "TOTAL"
        assert ws.cell(501, 4).value == "=SUM(D2:D500)"

        notas = wb["Notas"]
        assert notas.cell(1, 1).value == "Racional do teste"
        assert notas.cell(3, 1).value == "Segunda linha"

        # tipo inválido deve bloquear, não aceitar silenciosamente
        try:
            render({"parcelas": [{"tipo": "Multa", "valor": "1"}]}, template=TEMPLATE, output=Path(tmp) / "saida2.xlsx")
        except RenderError as exc:
            assert exc.code == "tipo_invalido", exc.code
        else:
            raise AssertionError("tipo de parcela inválido deveria ser bloqueado")

    print("ok — todos os checks passaram")


if __name__ == "__main__":
    run()
