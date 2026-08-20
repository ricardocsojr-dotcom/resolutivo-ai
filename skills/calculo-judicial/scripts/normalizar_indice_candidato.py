#!/usr/bin/env python3
"""Normalização local e passiva de fontes de índices.

O script aceita CSV, XLSX/XLSM e PDF. Ele não consulta a internet, não executa
macros, não recalcula planilhas e não altera o manifesto. A saída é um CSV
normalizado e um pacote JSON em status candidato.

Para XLS/XLSM, coluna de data, coluna de valor e linha inicial são obrigatórias.
Para PDF, um regex explícito com grupos nomeados ``data`` e ``valor`` é
obrigatório. O script não usa OCR nem adivinha a estrutura do PDF.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
import tempfile
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PT_MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

SERIES_TYPES = {
    "taxa_mensal_percentual",
    "taxa_diaria_decimal",
    "fator_acumulado",
    "taxa_aniversario_percentual",
    "numero_indice",
    "outro",
}
FREQUENCIES = {"mensal", "diaria", "por_aniversario", "outra"}


class NormalizeError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def error(code: str, message: str) -> NormalizeError:
    return NormalizeError(code, message)


def iso_date(value: Any, field: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str) or not value.strip():
        raise error("data_invalida", f"{field} deve conter uma data explícita.")
    raw = value.strip()
    try:
        return date.fromisoformat(raw)
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise error("data_invalida", f"{field} não pôde ser convertido sem inferência: {value!r}.")


def decimal(value: Any, field: str, *, decimal_comma: bool = False) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise error("valor_invalido", f"{field} deve conter um valor decimal explícito.")
    if isinstance(value, Decimal):
        result = value
    elif isinstance(value, (int, float)):
        result = Decimal(str(value))
    elif isinstance(value, str):
        raw = value.strip()
        if not raw:
            raise error("valor_ausente", f"{field} está vazio.")
        if "," in raw:
            if not decimal_comma:
                raise error("decimal_ambiguous", f"{field} usa vírgula decimal; declare --decimal-comma antes de normalizar.")
            raw = raw.replace(".", "").replace(",", ".")
        try:
            result = Decimal(raw)
        except InvalidOperation as exc:
            raise error("valor_invalido", f"{field} não é decimal válido: {value!r}.") from exc
    else:
        raise error("valor_invalido", f"{field} possui tipo não suportado.")
    if not result.is_finite():
        raise error("valor_nao_finito", f"{field} deve ser finito.")
    return result


def read_csv(path: Path, *, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise error("csv_nao_utf8", "CSV deve estar em UTF-8.") from exc
    reader = csv.DictReader(text.splitlines())
    if reader.fieldnames != ["data", "valor"]:
        raise error("cabecalho_invalido", "CSV deve conter exatamente o cabeçalho data,valor.")
    rows = []
    for line, record in enumerate(reader, start=2):
        if set(record) != {"data", "valor"}:
            raise error("colunas_invalidadas", f"Linha {line} contém colunas inesperadas.")
        rows.append((iso_date(record.get("data"), f"data da linha {line}"), decimal(record.get("valor"), f"valor da linha {line}", decimal_comma=decimal_comma)))
    return rows, {"formato": "csv", "macros_presentes": False, "linhas_fonte": len(rows)}


def column_number(letter: str) -> int:
    raw = letter.strip().upper()
    if not raw or not raw.isalpha():
        raise error("coluna_invalida", f"Coluna Excel inválida: {letter!r}.")
    value = 0
    for char in raw:
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def xlsx_has_vba(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as archive:
            return any(name.endswith("vbaProject.bin") for name in archive.namelist())
    except zipfile.BadZipFile as exc:
        raise error("arquivo_excel_invalido", f"Arquivo Excel inválido: {path}.") from exc


def read_xlsx_year_month(path: Path, *, sheet: str, year_column: str, month_column: str, value_column: str, start_row: int, end_row: int | None, month_language: str, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    if month_language != "pt-BR":
        raise error("idioma_mes_invalido", "Somente month-language=pt-BR está cadastrado.")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if sheet not in workbook.sheetnames:
        raise error("aba_inexistente", f"Aba não encontrada: {sheet!r}.")
    ws = workbook[sheet]
    year_col = column_number(year_column)
    month_col = column_number(month_column)
    value_col = column_number(value_column)
    last = end_row or ws.max_row
    rows: list[tuple[date, Decimal]] = []
    for row_number in range(start_row, last + 1):
        raw_year = ws.cell(row_number, year_col).value
        raw_month = ws.cell(row_number, month_col).value
        raw_value = ws.cell(row_number, value_col).value
        if raw_year in (None, "") and raw_month in (None, "") and raw_value in (None, ""):
            continue
        if raw_year in (None, "") or raw_month in (None, "") or raw_value in (None, ""):
            raise error("linha_incompleta", f"Linha {row_number} possui ano, mês ou valor ausente.")
        try:
            year = int(raw_year)
        except (TypeError, ValueError) as exc:
            raise error("ano_invalido", f"Ano inválido na linha {row_number}: {raw_year!r}.") from exc
        month_key = str(raw_month).strip().lower()
        month = PT_MONTHS.get(month_key)
        if month is None:
            raise error("mes_invalido", f"Mês em português não reconhecido na linha {row_number}: {raw_month!r}.")
        rows.append((date(year, month, 1), decimal(raw_value, f"valor da linha {row_number}", decimal_comma=decimal_comma)))
    workbook.close()
    return rows, {
        "formato": "xlsm" if path.suffix.lower() == ".xlsm" else "xlsx",
        "macros_presentes": xlsx_has_vba(path),
        "aba": sheet,
        "coluna_ano": year_column.upper(),
        "coluna_mes": month_column.upper(),
        "coluna_valor": value_column.upper(),
        "linha_inicial": start_row,
        "linha_final": last,
        "idioma_mes": month_language,
    }


def read_xlsx(path: Path, *, sheet: str, date_column: str, value_column: str, start_row: int, end_row: int | None, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    macros = xlsx_has_vba(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if sheet not in workbook.sheetnames:
        raise error("aba_inexistente", f"Aba não encontrada: {sheet!r}.")
    ws = workbook[sheet]
    date_col = column_number(date_column)
    value_col = column_number(value_column)
    last = end_row or ws.max_row
    if start_row < 1 or last < start_row:
        raise error("intervalo_invalido", "Intervalo de linhas inválido.")
    rows: list[tuple[date, Decimal]] = []
    for row_number in range(start_row, last + 1):
        raw_date = ws.cell(row_number, date_col).value
        raw_value = ws.cell(row_number, value_col).value
        if raw_date in (None, "") and raw_value in (None, ""):
            continue
        if raw_date in (None, "") or raw_value in (None, ""):
            raise error("linha_incompleta", f"Linha {row_number} possui apenas um dos campos.")
        rows.append((iso_date(raw_date, f"data da linha {row_number}"), decimal(raw_value, f"valor da linha {row_number}", decimal_comma=decimal_comma)))
    workbook.close()
    return rows, {
        "formato": "xlsm" if path.suffix.lower() == ".xlsm" else "xlsx",
        "macros_presentes": macros,
        "aba": sheet,
        "coluna_data": date_column.upper(),
        "coluna_valor": value_column.upper(),
        "linha_inicial": start_row,
        "linha_final": last,
    }


def read_legacy_xls_year_month(path: Path, *, sheet: str, year_column: str, month_column: str, value_column: str, start_row: int, end_row: int | None, month_language: str, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    try:
        with tempfile.TemporaryDirectory(prefix="rdaa-xls-") as tmp:
            tmp_path = Path(tmp)
            profile = (tmp_path / "profile").resolve()
            outdir = tmp_path / "out"
            outdir.mkdir()
            command = [
                "libreoffice", "--headless", "--norestore", "--nodefault", "--nofirststartwizard",
                f"-env:UserInstallation=file://{profile}", "--convert-to", "xlsx", "--outdir", str(outdir), str(path),
            ]
            subprocess.run(command, check=True, capture_output=True, text=True)
            converted = outdir / f"{path.stem}.xlsx"
            if not converted.is_file():
                raise error("xls_conversao_falhou", "LibreOffice não produziu XLSX para a fonte XLS.")
            rows, extraction = read_xlsx_year_month(
                converted,
                sheet=sheet,
                year_column=year_column,
                month_column=month_column,
                value_column=value_column,
                start_row=start_row,
                end_row=end_row,
                month_language=month_language,
                decimal_comma=decimal_comma,
            )
            extraction["formato"] = "xls_convertido_local"
            extraction["macros_presentes"] = "nao_verificado_no_conteiner_xls"
            extraction["conversao"] = "LibreOffice em perfil temporário, sem modo interativo"
            return rows, extraction
    except FileNotFoundError as exc:
        raise error("libreoffice_ausente", "LibreOffice não está disponível para converter XLS legado.") from exc
    except subprocess.CalledProcessError as exc:
        raise error("xls_conversao_falhou", f"Falha na conversão local do XLS: {exc.stderr.strip()}.") from exc


def read_legacy_xls(path: Path, *, sheet: str, date_column: str, value_column: str, start_row: int, end_row: int | None, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    """Converte XLS legado em área temporária para leitura somente dos valores.

    A conversão usa LibreOffice em perfil temporário e sem modo interativo.
    Nenhum macro é chamado pelo script. Como o contêiner XLS antigo não oferece
    inspeção segura de macros por esta biblioteca, o candidato registra essa
    limitação e permanece bloqueado.
    """
    try:
        with tempfile.TemporaryDirectory(prefix="rdaa-xls-") as tmp:
            tmp_path = Path(tmp)
            profile = (tmp_path / "profile").resolve()
            outdir = tmp_path / "out"
            outdir.mkdir()
            command = [
                "libreoffice",
                "--headless",
                "--norestore",
                "--nodefault",
                "--nofirststartwizard",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(outdir),
                str(path),
            ]
            subprocess.run(command, check=True, capture_output=True, text=True)
            converted = outdir / f"{path.stem}.xlsx"
            if not converted.is_file():
                raise error("xls_conversao_falhou", "LibreOffice não produziu XLSX para a fonte XLS.")
            rows, extraction = read_xlsx(
                converted,
                sheet=sheet,
                date_column=date_column,
                value_column=value_column,
                start_row=start_row,
                end_row=end_row,
                decimal_comma=decimal_comma,
            )
            extraction["formato"] = "xls_convertido_local"
            extraction["macros_presentes"] = "nao_verificado_no_conteiner_xls"
            extraction["conversao"] = "LibreOffice em perfil temporário, sem modo interativo"
            return rows, extraction
    except FileNotFoundError as exc:
        raise error("libreoffice_ausente", "LibreOffice não está disponível para converter XLS legado.") from exc
    except subprocess.CalledProcessError as exc:
        raise error("xls_conversao_falhou", f"Falha na conversão local do XLS: {exc.stderr.strip()}.") from exc


def read_pdf(path: Path, *, pattern: str, date_format: str, decimal_comma: bool) -> tuple[list[tuple[date, Decimal]], dict[str, Any]]:
    try:
        completed = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True)
    except FileNotFoundError as exc:
        raise error("pdftotext_ausente", "pdftotext não está disponível no ambiente local.") from exc
    except subprocess.CalledProcessError as exc:
        raise error("pdf_leitura_falhou", f"Não foi possível extrair texto do PDF: {exc.stderr.strip()}.") from exc
    try:
        matcher = re.compile(pattern, re.MULTILINE)
    except re.error as exc:
        raise error("regex_invalido", f"Regex de PDF inválido: {exc}.") from exc
    rows: list[tuple[date, Decimal]] = []
    for match in matcher.finditer(completed.stdout):
        groups = match.groupdict()
        if "data" not in groups or "valor" not in groups:
            raise error("regex_sem_grupos", "Regex de PDF deve ter grupos nomeados data e valor.")
        raw_date = groups["data"]
        try:
            record_date = datetime.strptime(raw_date, date_format).date()
        except ValueError as exc:
            raise error("data_pdf_invalida", f"Data de PDF inválida: {raw_date!r}.") from exc
        rows.append((record_date, decimal(groups["valor"].replace(" ", ""), f"valor de {record_date.isoformat()}", decimal_comma=decimal_comma)))
    return rows, {
        "formato": "pdf",
        "macros_presentes": False,
        "regex_declarado": pattern,
        "date_format": date_format,
        "matches": len(rows),
    }


def validate_rows(rows: Iterable[tuple[date, Decimal]]) -> list[tuple[date, Decimal]]:
    result = list(rows)
    if not result:
        raise error("sem_registros", "Nenhum registro foi extraído da fonte.")
    previous: date | None = None
    seen: set[date] = set()
    for record_date, _ in result:
        if record_date in seen:
            raise error("data_duplicada", f"Data duplicada na fonte: {record_date.isoformat()}.")
        if previous is not None and record_date <= previous:
            raise error("datas_fora_de_ordem", "Datas fora de ordem. O script não reordena a fonte silenciosamente.")
        seen.add(record_date)
        previous = record_date
    return result


def write_csv(path: Path, rows: list[tuple[date, Decimal]]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["data", "valor"])
        for record_date, value in rows:
            writer.writerow([record_date.isoformat(), format(value, "f")])
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build_candidate(args: argparse.Namespace) -> dict[str, Any]:
    if args.tipo_serie not in SERIES_TYPES:
        raise error("tipo_serie_invalido", args.tipo_serie)
    if args.frequencia not in FREQUENCIES:
        raise error("frequencia_invalida", args.frequencia)
    if not args.source.is_file():
        raise error("fonte_inexistente", f"Fonte não encontrada: {args.source}.")
    source_hash = hashlib.sha256(args.source.read_bytes()).hexdigest()
    suffix = args.source.suffix.lower()
    if suffix == ".csv":
        rows, extraction = read_csv(args.source, decimal_comma=args.decimal_comma)
    elif suffix in {".xlsx", ".xlsm"}:
        if not args.sheet or args.start_row is None:
            raise error("mapeamento_excel_ausente", "XLS/XLSM exige --sheet e --start-row.")
        if bool(args.year_column) != bool(args.month_column):
            raise error("mapeamento_ano_mes_incompleto", "year-column e month-column devem ser informados juntos.")
        if args.year_column and args.month_column:
            rows, extraction = read_xlsx_year_month(path=args.source, sheet=args.sheet, year_column=args.year_column, month_column=args.month_column, value_column=args.value_column, start_row=args.start_row, end_row=args.end_row, month_language=args.month_language, decimal_comma=args.decimal_comma)
        else:
            if not args.date_column or not args.value_column:
                raise error("mapeamento_excel_ausente", "XLS/XLSM exige date-column e value-column, ou year-column e month-column.")
            rows, extraction = read_xlsx(args.source, sheet=args.sheet, date_column=args.date_column, value_column=args.value_column, start_row=args.start_row, end_row=args.end_row, decimal_comma=args.decimal_comma)
    elif suffix == ".xls":
        if not args.allow_legacy_xls:
            raise error("xls_legado_exige_opt_in", "XLS legado exige --allow-legacy-xls para conversão local explícita.")
        if not args.sheet or args.start_row is None:
            raise error("mapeamento_excel_ausente", "XLS exige --sheet e --start-row.")
        if bool(args.year_column) != bool(args.month_column):
            raise error("mapeamento_ano_mes_incompleto", "year-column e month-column devem ser informados juntos.")
        if args.year_column and args.month_column:
            rows, extraction = read_legacy_xls_year_month(args.source, sheet=args.sheet, year_column=args.year_column, month_column=args.month_column, value_column=args.value_column, start_row=args.start_row, end_row=args.end_row, month_language=args.month_language, decimal_comma=args.decimal_comma)
        else:
            if not args.date_column or not args.value_column:
                raise error("mapeamento_excel_ausente", "XLS exige date-column e value-column, ou year-column e month-column.")
            rows, extraction = read_legacy_xls(args.source, sheet=args.sheet, date_column=args.date_column, value_column=args.value_column, start_row=args.start_row, end_row=args.end_row, decimal_comma=args.decimal_comma)
    elif suffix == ".pdf":
        if not args.pdf_regex:
            raise error("regex_pdf_ausente", "PDF exige --pdf-regex explícito com grupos data e valor.")
        rows, extraction = read_pdf(args.source, pattern=args.pdf_regex, date_format=args.pdf_date_format, decimal_comma=args.decimal_comma)
    else:
        raise error("formato_nao_suportado", f"Formato não suportado: {suffix}.")
    rows = validate_rows(rows)
    normalized_hash = write_csv(args.output_csv, rows)
    values = [value for _, value in rows]
    return {
        "schema_version": "1",
        "status": "candidato",
        "indice": args.indice,
        "definicao_proposta": {
            "arquivo": args.output_csv.name,
            "tipo_serie": args.tipo_serie,
            "unidade": args.unidade,
            "frequencia": args.frequencia,
            "convencoes": args.convencao,
        },
        "proveniencia": {
            "autoridade_primaria": args.autoridade,
            "url_ou_localizacao": args.url,
            "codigo_serie": args.codigo_serie,
            "data_coleta": args.data_coleta,
            "arquivo_bruto": str(args.source),
            "sha256_arquivo_bruto": source_hash,
            "observacoes": args.observacao,
        },
        "extracao": extraction,
        "integridade": {
            "sha256_csv_normalizado": normalized_hash,
            "registros": len(rows),
            "cobertura_inicio": rows[0][0].isoformat(),
            "cobertura_fim": rows[-1][0].isoformat(),
            "valor_minimo": format(min(values), "f"),
            "valor_maximo": format(max(values), "f"),
        },
        "bloqueio": "Não altera o manifesto. Exige caso dourado e aprovação explícita.",
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Normalizar fonte local de índice em candidato auditável")
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--indice", required=True)
    parser.add_argument("--tipo-serie", required=True, choices=sorted(SERIES_TYPES))
    parser.add_argument("--unidade", required=True)
    parser.add_argument("--frequencia", required=True, choices=sorted(FREQUENCIES))
    parser.add_argument("--convencao", action="append", required=True)
    parser.add_argument("--autoridade", required=True)
    parser.add_argument("--url", required=True)
    parser.add_argument("--codigo-serie", default="")
    parser.add_argument("--data-coleta", default="")
    parser.add_argument("--observacao", default="")
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--output-json", required=True, type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--date-column")
    parser.add_argument("--year-column")
    parser.add_argument("--month-column")
    parser.add_argument("--month-language", default="pt-BR")
    parser.add_argument("--value-column")
    parser.add_argument("--start-row", type=int)
    parser.add_argument("--end-row", type=int)
    parser.add_argument("--allow-legacy-xls", action="store_true")
    parser.add_argument("--decimal-comma", action="store_true")
    parser.add_argument("--pdf-regex")
    parser.add_argument("--pdf-date-format", default="%d/%m/%Y")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        candidate = build_candidate(args)
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(json.dumps(candidate, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(json.dumps({"status": "ok", "saida_csv": str(args.output_csv), "saida_json": str(args.output_json), "sha256_csv": candidate["integridade"]["sha256_csv_normalizado"]}, ensure_ascii=False))
        return 0
    except (OSError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "erro", "codigo": "entrada_arquivo_invalida", "mensagem": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except NormalizeError as exc:
        print(json.dumps({"status": "erro", "codigo": exc.code, "mensagem": exc.message}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
