#!/usr/bin/env python3
"""Renderiza um pacote de memória em uma cópia do template RDAA.

O renderizador não calcula valores e não escolhe regras. Ele apenas coloca em
uma planilha os parâmetros e resultados que já foram declarados e produzidos
pelo motor ou conferidos pelo responsável.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STATUSES = {"candidato", "aprovado", "bloqueado"}


class RenderError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def error(code: str, message: str) -> RenderError:
    return RenderError(code, message)


def obj(value: Any, field: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise error("objeto_invalido", f"{field} deve ser objeto JSON.")
    return value


def text(value: Any, field: str, *, required: bool = False) -> str:
    if value is None:
        if required:
            raise error("campo_ausente", f"{field} é obrigatório.")
        return ""
    if not isinstance(value, str):
        raise error("texto_invalido", f"{field} deve ser texto explícito.")
    if required and not value.strip():
        raise error("campo_ausente", f"{field} é obrigatório.")
    return value


def money(value: Any, field: str, *, required: bool = False) -> Decimal | None:
    raw = text(value, field, required=required)
    if raw == "":
        return None
    try:
        result = Decimal(raw)
    except InvalidOperation as exc:
        raise error("valor_monetario_invalido", f"{field} não é decimal válido.") from exc
    if not result.is_finite():
        raise error("valor_monetario_invalido", f"{field} deve ser finito.")
    return result


def set_if_present(ws, cell: str, value: Any) -> None:
    if value is not None and value != "":
        ws[cell] = value


def render(payload: Any, *, template: Path, output: Path) -> None:
    data = obj(payload, "entrada")
    lancamentos = data.get("lancamentos")
    if not isinstance(lancamentos, list) or not lancamentos:
        raise error("lancamentos_ausentes", "entrada.lancamentos deve ser lista não vazia.")
    if len(lancamentos) > 200:
        raise error("lancamentos_excedentes", "O template suporta no máximo 200 lançamentos.")

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, output)
    workbook = load_workbook(output, read_only=False, data_only=False, keep_vba=False)
    summary = workbook["Resumo"]
    entries = workbook["Lançamentos"]
    indices = workbook["Índices"]
    cases = workbook["Casos dourados"]
    rules = workbook["Regras declaradas"]
    interest_segments_sheet = workbook["Segmentos de juros"]

    meta = obj(data.get("meta", {}), "meta")
    set_if_present(summary, "B3", text(meta.get("materia"), "meta.materia"))
    set_if_present(summary, "B4", text(meta.get("processo"), "meta.processo"))
    set_if_present(summary, "B5", text(meta.get("tribunal_uf"), "meta.tribunal_uf"))
    set_if_present(summary, "B6", text(meta.get("data_final"), "meta.data_final"))
    set_if_present(summary, "B7", text(meta.get("status", "candidato"), "meta.status"))
    if summary["B7"].value not in STATUSES:
        raise error("status_invalido", "meta.status deve ser candidato, aprovado ou bloqueado.")

    for row_index, raw_entry in enumerate(lancamentos, start=3):
        entry = obj(raw_entry, f"lancamentos[{row_index - 3}]")
        status = text(entry.get("status", "candidato"), f"lancamentos[{row_index - 3}].status")
        if status not in STATUSES:
            raise error("status_invalido", f"Status inválido no lançamento {row_index - 3}.")
        values = {
            1: text(entry.get("id"), f"lancamentos[{row_index - 3}].id", required=True),
            2: text(entry.get("grupo"), f"lancamentos[{row_index - 3}].grupo"),
            3: text(entry.get("descricao"), f"lancamentos[{row_index - 3}].descricao"),
            4: text(entry.get("data_base"), f"lancamentos[{row_index - 3}].data_base"),
            5: money(entry.get("valor_original"), f"lancamentos[{row_index - 3}].valor_original", required=True),
            6: text(entry.get("fonte_indice"), f"lancamentos[{row_index - 3}].fonte_indice"),
            7: text(entry.get("unidade_frequencia"), f"lancamentos[{row_index - 3}].unidade_frequencia"),
            8: money(entry.get("fator_autorizado"), f"lancamentos[{row_index - 3}].fator_autorizado"),
            9: money(entry.get("valor_corrigido"), f"lancamentos[{row_index - 3}].valor_corrigido"),
            10: text(entry.get("inicio_juros"), f"lancamentos[{row_index - 3}].inicio_juros"),
            11: text(entry.get("fim_juros"), f"lancamentos[{row_index - 3}].fim_juros"),
            12: text(entry.get("convencao_juros"), f"lancamentos[{row_index - 3}].convencao_juros"),
            13: money(entry.get("juros"), f"lancamentos[{row_index - 3}].juros"),
            14: money(entry.get("multa_percentual"), f"lancamentos[{row_index - 3}].multa_percentual"),
            15: money(entry.get("multa"), f"lancamentos[{row_index - 3}].multa"),
            16: money(entry.get("honorarios_percentual"), f"lancamentos[{row_index - 3}].honorarios_percentual"),
            17: money(entry.get("honorarios"), f"lancamentos[{row_index - 3}].honorarios"),
            18: money(entry.get("custas"), f"lancamentos[{row_index - 3}].custas"),
            20: status,
            21: text(entry.get("fonte_caso"), f"lancamentos[{row_index - 3}].fonte_caso"),
            22: text(entry.get("observacoes"), f"lancamentos[{row_index - 3}].observacoes"),
        }
        for column, value in values.items():
            if value is not None and value != "":
                entries.cell(row_index, column, value)
        total = money(entry.get("total"), f"lancamentos[{row_index - 3}].total")
        if total is not None:
            entries.cell(row_index, 19, total)
        segments = entry.get("segmentos_juros", [])
        if not isinstance(segments, list):
            raise error("segmentos_invalidos", f"lancamentos[{row_index - 3}].segmentos_juros deve ser lista.")
        for segment_index, raw_segment in enumerate(segments, start=3):
            segment = obj(raw_segment, f"lancamentos[{row_index - 3}].segmentos_juros[{segment_index - 3}]")
            segment_values = [
                text(entry.get("id"), f"lancamentos[{row_index - 3}].id", required=True),
                segment_index - 2,
                text(segment.get("inicio"), "segmento.inicio"),
                text(segment.get("fim"), "segmento.fim"),
                money(segment.get("taxa"), "segmento.taxa"),
                text(segment.get("unidade_taxa"), "segmento.unidade_taxa"),
                text(segment.get("base"), "segmento.base"),
                text(segment.get("convencao"), "segmento.convencao"),
                text(segment.get("meses"), "segmento.meses"),
                money(segment.get("juros"), "segmento.juros"),
                text(segment.get("status", entry.get("status", "candidato")), "segmento.status"),
            ]
            if segment_values[10] not in {"candidato", "aprovado", "bloqueado"}:
                raise error("status_invalido", "Status de segmento deve ser candidato, aprovado ou bloqueado.")
            for column, value in enumerate(segment_values, 1):
                if value is not None and value != "":
                    interest_segments_sheet.cell(segment_index, column, value)

    index_info = data.get("indice")
    if index_info is not None:
        index = obj(index_info, "indice")
        row = 3
        index_values = [
            text(index.get("id"), "indice.id", required=True),
            text(index.get("nome"), "indice.nome"),
            text(index.get("autoridade_primaria"), "indice.autoridade_primaria"),
            text(index.get("url_ou_localizacao"), "indice.url_ou_localizacao"),
            text(index.get("codigo_serie"), "indice.codigo_serie"),
            text(index.get("arquivo_bruto"), "indice.arquivo_bruto"),
            text(index.get("sha256_arquivo_bruto"), "indice.sha256_arquivo_bruto"),
            text(index.get("csv_normalizado"), "indice.csv_normalizado"),
            text(index.get("sha256_csv_normalizado"), "indice.sha256_csv_normalizado"),
            text(index.get("tipo_serie"), "indice.tipo_serie"),
            text(index.get("unidade"), "indice.unidade"),
            text(index.get("frequencia"), "indice.frequencia"),
            text(index.get("cobertura_inicio"), "indice.cobertura_inicio"),
            text(index.get("cobertura_fim"), "indice.cobertura_fim"),
            text(index.get("status", "candidato"), "indice.status"),
            text(index.get("observacoes"), "indice.observacoes"),
        ]
        if index_values[14] not in STATUSES:
            raise error("status_invalido", "indice.status deve ser candidato, aprovado ou bloqueado.")
        for column, value in enumerate(index_values, 1):
            indices.cell(row, column, value)

    for row_index, raw_case in enumerate(data.get("casos_dourados", []), start=3):
        case = obj(raw_case, f"casos_dourados[{row_index - 3}]")
        values = [
            text(case.get("caso"), f"casos_dourados[{row_index - 3}].caso"),
            text(case.get("indice"), f"casos_dourados[{row_index - 3}].indice"),
            text(case.get("entrada_resumida"), f"casos_dourados[{row_index - 3}].entrada_resumida"),
            text(case.get("convencao_declarada"), f"casos_dourados[{row_index - 3}].convencao_declarada"),
            text(case.get("resultado_esperado"), f"casos_dourados[{row_index - 3}].resultado_esperado"),
            text(case.get("tolerancia"), f"casos_dourados[{row_index - 3}].tolerancia"),
            text(case.get("memoria_evidencia"), f"casos_dourados[{row_index - 3}].memoria_evidencia"),
            text(case.get("status", "candidato"), f"casos_dourados[{row_index - 3}].status"),
            text(case.get("aprovado_por"), f"casos_dourados[{row_index - 3}].aprovado_por"),
            text(case.get("data_observacoes"), f"casos_dourados[{row_index - 3}].data_observacoes"),
        ]
        if values[7] not in STATUSES:
            raise error("status_invalido", f"Status inválido no caso dourado {row_index - 3}.")
        for column, value in enumerate(values, 1):
            cases.cell(row_index, column, value)

    for row_index, raw_rule in enumerate(data.get("regras", []), start=3):
        rule = obj(raw_rule, f"regras[{row_index - 3}]")
        values = [
            text(rule.get("componente"), f"regras[{row_index - 3}].componente"),
            text(rule.get("regra"), f"regras[{row_index - 3}].regra"),
            text(rule.get("fonte_declaracao"), f"regras[{row_index - 3}].fonte_declaracao"),
            text(rule.get("status", "candidato"), f"regras[{row_index - 3}].status"),
            text(rule.get("efeito"), f"regras[{row_index - 3}].efeito"),
            text(rule.get("observacao"), f"regras[{row_index - 3}].observacao"),
        ]
        if values[3] not in STATUSES:
            raise error("status_invalido", f"Status inválido na regra {row_index - 3}.")
        for column, value in enumerate(values, 1):
            rules.cell(row_index, column, value)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Renderizar memória JSON no template RDAA")
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--template", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args(argv)
    try:
        payload = json.loads(args.input.read_text(encoding="utf-8"))
        render(payload, template=args.template, output=args.output)
        print(json.dumps({"status": "ok", "saida": str(args.output)}, ensure_ascii=False))
        return 0
    except (OSError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "erro", "codigo": "entrada_arquivo_invalida", "mensagem": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except RenderError as exc:
        print(json.dumps({"status": "erro", "codigo": exc.code, "mensagem": exc.message}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
