#!/usr/bin/env python3
"""Renderiza uma memória de cálculo simples (uma linha por parcela, tabela
única) no template-calculo-simples-rdaa.xlsx.

Não calcula valor nenhum — só coloca em planilha o que já foi produzido
pelo motor ou declarado por quem pediu o cálculo. Uso: casos com um único
índice de correção e sem juros segmentados/múltiplos lançamentos
complexos. Para isso, ver renderizar_memoria_template.py.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TIPOS = {"Principal", "Honorários", "Custas"}
COLUNAS = (
    "tipo", "mes_ref", "vencimento", "valor",
    "fator_multa", "multa",
    "fator_correcao", "correcao",
    "fator_juros", "juros",
    "fator_encargos", "encargos",
    "subtotal",
)


class RenderError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def error(code: str, message: str) -> RenderError:
    return RenderError(code, message)


def texto(value: Any, campo: str) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        raise error("texto_invalido", f"{campo} deve ser texto.")
    return value


def numero(value: Any, campo: str):
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
    except InvalidOperation as exc:
        raise error("valor_invalido", f"{campo} deve ser decimal explícito.") from exc
    if not d.is_finite():
        raise error("valor_invalido", f"{campo} deve ser finito.")
    return float(d)


def render(payload: dict[str, Any], *, template: Path, output: Path) -> int:
    if not isinstance(payload, dict):
        raise error("entrada_invalida", "Entrada deve ser um objeto JSON.")
    parcelas = payload.get("parcelas")
    if not isinstance(parcelas, list) or not parcelas:
        raise error("parcelas_ausentes", "parcelas deve ser lista não vazia.")
    if len(parcelas) > 499:
        raise error("parcelas_excedentes", "O template suporta no máximo 499 parcelas.")

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, output)
    wb = load_workbook(output)
    ws = wb["Cálculo"]

    for row_index, raw in enumerate(parcelas, start=2):
        if not isinstance(raw, dict):
            raise error("parcela_invalida", f"parcelas[{row_index - 2}] deve ser objeto.")
        tipo = texto(raw.get("tipo"), f"parcelas[{row_index - 2}].tipo")
        if tipo and tipo not in TIPOS:
            raise error("tipo_invalido", f"parcelas[{row_index - 2}].tipo deve ser um de {sorted(TIPOS)}.")
        valores = [
            tipo,
            texto(raw.get("mes_ref"), "mes_ref"),
            texto(raw.get("vencimento"), "vencimento"),
            numero(raw.get("valor"), "valor"),
            numero(raw.get("fator_multa"), "fator_multa"),
            numero(raw.get("multa"), "multa"),
            numero(raw.get("fator_correcao"), "fator_correcao"),
            numero(raw.get("correcao"), "correcao"),
            numero(raw.get("fator_juros"), "fator_juros"),
            numero(raw.get("juros"), "juros"),
            numero(raw.get("fator_encargos"), "fator_encargos"),
            numero(raw.get("encargos"), "encargos"),
            numero(raw.get("subtotal"), "subtotal"),
        ]
        for col_index, value in enumerate(valores, start=1):
            if value not in (None, ""):
                ws.cell(row_index, col_index, value)

    notas = payload.get("notas")
    if notas is not None:
        if isinstance(notas, str):
            notas = [notas]
        if not isinstance(notas, list) or not all(isinstance(item, str) for item in notas):
            raise error("notas_invalidas", "notas deve ser texto ou lista de textos.")
        notas_ws = wb["Notas"]
        for row_index, linha in enumerate(notas, start=1):
            notas_ws.cell(row_index, 1, linha)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    return len(parcelas)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Renderizar memória simples (parcela única por linha) no template RDAA")
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--template", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args(argv)
    try:
        payload = json.loads(args.input.read_text(encoding="utf-8"))
        total = render(payload, template=args.template, output=args.output)
        print(json.dumps({"status": "ok", "saida": str(args.output), "parcelas": total}, ensure_ascii=False))
        return 0
    except (OSError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "erro", "codigo": "entrada_arquivo_invalida", "mensagem": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    except RenderError as exc:
        print(json.dumps({"status": "erro", "codigo": exc.code, "mensagem": exc.message}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
