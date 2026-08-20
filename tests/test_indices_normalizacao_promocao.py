from __future__ import annotations

import hashlib
import importlib.util
import json
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
NORMALIZER = ROOT / "skills" / "calculo-judicial" / "scripts" / "normalizar_indice_candidato.py"
PROMOTER = ROOT / "skills" / "calculo-judicial" / "scripts" / "promover_indice_aprovado.py"
PREPARER = ROOT / "skills" / "calculo-judicial" / "scripts" / "preparar_fonte_candidata.py"


def _load(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_preparer():
    return _load(PREPARER, "preparar_fonte_candidata")


def _base_args(module, source: Path, output_csv: Path, output_json: Path):
    return module.build_parser().parse_args(
        [
            "--source", str(source),
            "--indice", "teste",
            "--tipo-serie", "taxa_mensal_percentual",
            "--unidade", "percentual_mensal",
            "--frequencia", "mensal",
            "--convencao", "meses_calendario_inclusivos",
            "--autoridade", "Fonte oficial de teste",
            "--url", "arquivo local",
            "--codigo-serie", "T-1",
            "--data-coleta", "2026-08-20",
            "--output-csv", str(output_csv),
            "--output-json", str(output_json),
        ]
    )


def test_normalizer_csv_requires_explicit_decimal_comma() -> None:
    module = _load(NORMALIZER, "normalizar_indice_candidato")
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "fonte.csv"
        source.write_text("data,valor\n2024-01-01,1,25\n", encoding="utf-8")
        args = _base_args(module, source, root / "out.csv", root / "candidate.json")
        try:
            module.build_candidate(args)
        except module.NormalizeError as exc:
            assert exc.code == "colunas_invalidadas"
        else:
            raise AssertionError("CSV ambíguo deveria bloquear")


def test_normalizer_xlsx_requires_explicit_mapping() -> None:
    module = _load(NORMALIZER, "normalizar_indice_candidato")
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "fonte.xlsx"
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Índices"
        ws["A2"] = "2024-01-01"
        ws["B2"] = 1.25
        workbook.save(source)
        args = _base_args(module, source, root / "out.csv", root / "candidate.json")
        args.sheet = "Índices"
        args.date_column = "A"
        args.value_column = "B"
        args.start_row = 2
        candidate = module.build_candidate(args)
        assert candidate["status"] == "candidato"
        assert candidate["extracao"]["aba"] == "Índices"
        assert (root / "out.csv").read_text(encoding="utf-8").splitlines()[1] == "2024-01-01,1.25"


def test_preparer_uses_tjsp_profile_without_manifest_change() -> None:
    module = _load_preparer()
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "tjsp.xlsx"
        workbook = Workbook()
        ws = workbook.active
        ws.title = "ÍNDICES"
        ws["A7"] = "2024-01-01"
        ws["O7"] = 1.25
        workbook.save(source)
        output_dir = root / "candidate"
        args = module.parser().parse_args([
            "--profile", "tjsp-planilha-indices",
            "--source", str(source),
            "--indice", "tjsp-candidato-teste",
            "--output-dir", str(output_dir),
            "--data-coleta", "2026-08-20",
        ])
        candidate = module.prepare(args)
        assert candidate["status"] == "candidato"
        assert (output_dir / "tjsp-candidato-teste.csv").is_file()
        assert json.loads((output_dir / "tjsp-candidato-teste.candidate.json").read_text(encoding="utf-8"))["status"] == "candidato"


def test_normalizer_xlsx_year_month_profile() -> None:
    module = _load(NORMALIZER, "normalizar_indice_candidato")
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "tjmg.xlsx"
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Plan1"
        ws["A2"] = 2026
        ws["B2"] = "Julho"
        ws["C2"] = 0.9999
        workbook.save(source)
        args = _base_args(module, source, root / "out.csv", root / "candidate.json")
        args.sheet = "Plan1"
        args.year_column = "A"
        args.month_column = "B"
        args.month_language = "pt-BR"
        args.value_column = "C"
        args.start_row = 2
        candidate = module.build_candidate(args)
        assert candidate["integridade"]["cobertura_inicio"] == "2026-07-01"
        assert (root / "out.csv").read_text(encoding="utf-8").splitlines()[1] == "2026-07-01,0.9999"


def test_normalizer_pdf_requires_declared_regex_and_preserves_order() -> None:
    module = _load(NORMALIZER, "normalizar_indice_candidato")
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        source = root / "fonte.pdf"
        source.write_bytes(b"PDF controlado")
        args = _base_args(module, source, root / "out.csv", root / "candidate.json")
        args.pdf_regex = r"(?P<data>\d{2}/\d{2}/\d{4})\s+(?P<valor>[0-9,]+)"
        args.decimal_comma = True
        with patch.object(module.subprocess, "run", return_value=type("Completed", (), {"stdout": "01/01/2024 1,25\n01/02/2024 2,50\n"})()):
            candidate = module.build_candidate(args)
        assert candidate["extracao"]["formato"] == "pdf"
        assert candidate["integridade"]["registros"] == 2


def test_promotion_writes_copy_and_preserves_source_manifest() -> None:
    module = _load(PROMOTER, "promover_indice_aprovado")
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        csv = root / "teste.csv"
        csv.write_text("data,valor\n2024-01-01,1.25\n", encoding="utf-8")
        digest = hashlib.sha256(csv.read_bytes()).hexdigest()
        manifest = root / "manifest.json"
        manifest.write_text(json.dumps({"schema_version": "1", "indices": {"teste": {"status": "candidato"}}}), encoding="utf-8")
        candidate = root / "candidate.json"
        candidate.write_text(json.dumps({
            "status": "candidato",
            "indice": "teste",
            "definicao_proposta": {
                "arquivo": "teste.csv",
                "tipo_serie": "taxa_mensal_percentual",
                "unidade": "percentual_mensal",
                "frequencia": "mensal",
                "convencoes": ["meses_calendario_inclusivos"],
            },
            "proveniencia": {
                "autoridade_primaria": "Fonte",
                "url_ou_localizacao": "local",
                "codigo_serie": "T-1",
                "data_coleta": "2026-08-20",
                "arquivo_bruto": "bruto.xlsx",
                "sha256_arquivo_bruto": "a" * 64,
            },
            "integridade": {"sha256_csv_normalizado": digest},
        }), encoding="utf-8")
        golden = root / "golden.json"
        golden.write_text(json.dumps({
            "status": "aprovado",
            "caso": "T-1",
            "formula_referencia": "fator declarado",
            "convencao_declarada": "meses inclusivos",
            "resultado_esperado": "100.00",
            "resultado_observado": "100.00",
            "tolerancia": "0.01",
            "aprovado_por": "Ricardo",
            "data_aprovacao": "2026-08-20",
        }), encoding="utf-8")
        output = root / "manifest-promoted.json"
        before = manifest.read_bytes()
        module.promote(manifest_path=manifest, candidate_path=candidate, normalized_csv=csv, golden_path=golden, output_manifest=output)
        assert manifest.read_bytes() == before
        promoted = json.loads(output.read_text(encoding="utf-8"))
        assert promoted["indices"]["teste"]["status"] == "aprovado"
        assert promoted["indices"]["teste"]["sha256"] == digest


def main() -> None:
    test_normalizer_csv_requires_explicit_decimal_comma()
    test_normalizer_xlsx_requires_explicit_mapping()
    test_preparer_uses_tjsp_profile_without_manifest_change()
    test_normalizer_xlsx_year_month_profile()
    test_normalizer_pdf_requires_declared_regex_and_preserves_order()
    test_promotion_writes_copy_and_preserves_source_manifest()
    print("PASS test_indices_normalizacao_promocao")


if __name__ == "__main__":
    main()
