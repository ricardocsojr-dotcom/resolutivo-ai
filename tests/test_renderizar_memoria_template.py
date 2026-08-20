from __future__ import annotations

import importlib.util
import json
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "calculo-judicial" / "scripts" / "renderizar_memoria_template.py"
TEMPLATE = ROOT / "skills" / "calculo-judicial" / "references" / "template-calculo-rdaa.xlsx"


def _load_module():
    spec = importlib.util.spec_from_file_location("renderizar_memoria_template", SCRIPT)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _payload(status: str = "candidato") -> dict:
    return {
        "meta": {
            "materia": "Teste de memória",
            "processo": "0000000-00.0000.0.00.0000",
            "tribunal_uf": "TJMG / MG",
            "data_final": "2026-08-20",
            "status": status,
        },
        "indice": {
            "id": "teste",
            "nome": "Índice de teste",
            "autoridade_primaria": "Fonte declarada",
            "url_ou_localizacao": "arquivo local",
            "codigo_serie": "T-1",
            "arquivo_bruto": "bruto.xlsx",
            "sha256_arquivo_bruto": "a" * 64,
            "csv_normalizado": "teste.csv",
            "sha256_csv_normalizado": "b" * 64,
            "tipo_serie": "fator acumulado",
            "unidade": "fator",
            "frequencia": "mensal",
            "cobertura_inicio": "2024-01-01",
            "cobertura_fim": "2026-08-01",
            "status": status,
            "observacoes": "caso controlado",
        },
        "lancamentos": [
            {
                "id": "L-1",
                "grupo": "Condenação",
                "descricao": "Item de teste",
                "data_base": "2024-01-01",
                "valor_original": "1000.00",
                "fonte_indice": "teste",
                "unidade_frequencia": "fator acumulado / mensal",
                "fator_autorizado": "1.20",
                "valor_corrigido": "1200.00",
                "inicio_juros": "2024-02-01",
                "fim_juros": "2026-08-20",
                "convencao_juros": "declarada",
                "juros": "100.00",
                "multa_percentual": "0",
                "multa": "0.00",
                "honorarios_percentual": "10",
                "honorarios": "120.00",
                "custas": "0.00",
                "total": "1420.00",
                "status": status,
                "fonte_caso": "caso dourado T-1",
                "observacoes": "sem inferência",
                "segmentos_juros": [
                    {
                        "inicio": "2024-02-01",
                        "fim": "2026-08-31",
                        "taxa": "1.00",
                        "unidade_taxa": "percentual_mensal",
                        "base": "principal",
                        "convencao": "meses_calendario_inclusivos",
                        "meses": "31",
                        "juros": "100.00",
                        "status": status,
                    }
                ],
            }
        ],
        "casos_dourados": [
            {
                "caso": "T-1",
                "indice": "teste",
                "entrada_resumida": "1000 de 2024-01 a 2026-08",
                "convencao_declarada": "meses inclusivos",
                "resultado_esperado": "1420.00",
                "tolerancia": "0.01",
                "memoria_evidencia": "conferido manualmente",
                "status": status,
                "aprovado_por": "",
                "data_observacoes": "",
            }
        ],
        "regras": [
            {
                "componente": "Juros",
                "regra": "1% mensal conforme caso controlado",
                "fonte_declaracao": "documento de teste",
                "status": status,
                "efeito": "Aplicar somente se aprovado",
                "observacao": "",
            }
        ],
    }


def test_render_copies_template_and_writes_explicit_values() -> None:
    module = _load_module()
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        output = root / "memory.xlsx"
        module.render(_payload(), template=TEMPLATE, output=output)
        workbook = load_workbook(output, data_only=False)
        assert workbook.sheetnames == [
            "Instruções",
            "Resumo",
            "Lançamentos",
            "Segmentos de juros",
            "Índices",
            "Casos dourados",
            "Regras declaradas",
        ]
        assert workbook["Resumo"]["B4"].value == "0000000-00.0000.0.00.0000"
        assert workbook["Lançamentos"]["A3"].value == "L-1"
        assert float(workbook["Lançamentos"]["I3"].value) == 1200.00
        assert float(workbook["Lançamentos"]["S3"].value) == 1420.00
        assert workbook["Lançamentos"]["T3"].value == "candidato"
        assert workbook["Segmentos de juros"]["A3"].value == "L-1"
        assert float(workbook["Segmentos de juros"]["E3"].value) == 1.00
        assert workbook["Índices"]["A3"].value == "teste"
        assert workbook["Casos dourados"]["A3"].value == "T-1"


def test_render_rejects_non_explicit_status() -> None:
    module = _load_module()
    with TemporaryDirectory() as tmp:
        output = Path(tmp) / "memory.xlsx"
        payload = _payload(status="inferido")
        try:
            module.render(payload, template=TEMPLATE, output=output)
        except module.RenderError as exc:
            assert exc.code == "status_invalido"
        else:
            raise AssertionError("status não declarado deveria bloquear a renderização")


def main() -> None:
    test_render_copies_template_and_writes_explicit_values()
    test_render_rejects_non_explicit_status()
    print("PASS test_renderizar_memoria_template")


if __name__ == "__main__":
    main()
