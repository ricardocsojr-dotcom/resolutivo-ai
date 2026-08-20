from __future__ import annotations

import hashlib
import importlib.util
import json
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "calculo-judicial" / "scripts" / "registrar_indice_candidato.py"
TEMPLATE = ROOT / "skills" / "calculo-judicial" / "references" / "template-calculo-rdaa.xlsx"
MANIFEST = ROOT / "skills" / "calculo-judicial" / "references" / "index_manifest.json"


def _load_script():
    spec = importlib.util.spec_from_file_location("registrar_indice_candidato", SCRIPT)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_csv(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def test_candidate_is_pending_and_manifest_is_untouched() -> None:
    module = _load_script()
    with TemporaryDirectory() as tmp:
        root = Path(tmp)
        csv_path = root / "teste.csv"
        output_path = root / "candidate.json"
        _write_csv(csv_path, "data,valor\n2024-01-01,1.00\n2024-02-01,2.00\n")
        before = MANIFEST.read_bytes()
        args = module.parser().parse_args(
            [
                "--csv", str(csv_path),
                "--indice", "teste",
                "--tipo-serie", "taxa_mensal_percentual",
                "--unidade", "percentual_mensal",
                "--frequencia", "mensal",
                "--convencao", "meses_calendario_inclusivos",
                "--autoridade", "Fonte teste",
                "--url", "https://example.invalid/fonte",
                "--codigo-serie", "ABC",
                "--data-coleta", "2026-08-20",
                "--output", str(output_path),
            ]
        )
        candidate = module.build_candidate(args)
        output_path.write_text(json.dumps(candidate, ensure_ascii=False, indent=2), encoding="utf-8")
        data = json.loads(output_path.read_text(encoding="utf-8"))
        assert data["status"] == "candidato"
        assert data["integridade"]["registros"] == 2
        assert data["integridade"]["sha256_csv"] == hashlib.sha256(csv_path.read_bytes()).hexdigest()
        assert data["bloqueio"]
        assert MANIFEST.read_bytes() == before


def test_candidate_rejects_duplicate_dates() -> None:
    module = _load_script()
    with TemporaryDirectory() as tmp:
        path = Path(tmp) / "duplicado.csv"
        _write_csv(path, "data,valor\n2024-01-01,1\n2024-01-01,2\n")
        try:
            module.read_csv(path)
        except module.CandidateError as exc:
            assert exc.code == "data_duplicada"
        else:
            raise AssertionError("data duplicada deveria bloquear o candidato")


def test_candidate_rejects_nonfinite_values() -> None:
    module = _load_script()
    with TemporaryDirectory() as tmp:
        path = Path(tmp) / "infinito.csv"
        _write_csv(path, "data,valor\n2024-01-01,NaN\n")
        try:
            module.read_csv(path)
        except module.CandidateError as exc:
            assert exc.code == "valor_nao_finito"
        else:
            raise AssertionError("valor não finito deveria bloquear o candidato")


def test_template_has_required_sheets_and_no_macro() -> None:
    assert TEMPLATE.is_file()
    with TEMPLATE.open("rb") as handle:
        header = handle.read(4)
    assert header == b"PK\x03\x04"
    workbook = load_workbook(TEMPLATE, read_only=False, data_only=False, keep_vba=False)
    assert workbook.sheetnames == [
        "Instruções",
        "Resumo",
        "Lançamentos",
        "Segmentos de juros",
        "Índices",
        "Casos dourados",
        "Regras declaradas",
    ]
    assert workbook["Lançamentos"]["T3"].value in (None, "")
    assert workbook["Lançamentos"]["S3"].value.startswith("=IF(")
    assert workbook["Índices"]["O3"].value == "candidato"
    assert workbook["Casos dourados"]["H3"].value in (None, "")
    assert not getattr(workbook, "_external_links", [])


def main() -> None:
    test_candidate_is_pending_and_manifest_is_untouched()
    test_candidate_rejects_duplicate_dates()
    test_candidate_rejects_nonfinite_values()
    test_template_has_required_sheets_and_no_macro()
    print("PASS test_template_calculo_rdaa")


if __name__ == "__main__":
    main()
